"""카데이터 팀 (물로간다) 작업 현황 Excel 파일 생성 스크립트.

API 키, 서비스 계정 등 사전 준비 없이 바로 실행할 수 있다.

의존성 설치
----------
    pip install openpyxl   (requirements.txt 에 이미 포함)

실행
----
    python ai_tools/google_sheet_setup.py

실행 결과
---------
    물로간다_팀작업현황.xlsx  파일이 프로젝트 루트에 생성된다.

구글 스프레드시트로 올리기 (3단계)
-----------------------------------
1. https://drive.google.com 접속 → '새로 만들기' → '파일 업로드'
   → 생성된 xlsx 파일 업로드
2. 업로드된 파일 우클릭 → 'Google 스프레드시트로 열기'
3. 공유 버튼 → '링크가 있는 모든 사용자' → '편집자' → 링크 복사
   → 팀 단톡방에 링크 공유
"""

from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
except ImportError:
    print("[오류] openpyxl 이 설치되어 있지 않습니다.\n       pip install openpyxl")
    sys.exit(1)


# ── 출력 파일 경로 ─────────────────────────────────────────────────────
OUTPUT_PATH = Path(__file__).parent.parent / "물로간다_팀작업현황.xlsx"

# ── 드롭다운 선택지 ────────────────────────────────────────────────────
TASK_STATUS_LIST  = '"Todo,Doing,Done,Blocked"'
ISSUE_STATUS_LIST = '"Open,Closed"'
MEMBER_LIST       = '"최연우,권소라,김지혜,박회종,전원"'
AREA_LIST         = '"DB,Crawler,Backend,UI,Docs"'

# ── 상태별 조건부 서식 색상 ────────────────────────────────────────────
TASK_CF_RULES = [
    ("Todo",    "EBF5FB"),   # 연파랑
    ("Doing",   "FFEB9C"),   # 노랑
    ("Done",    "C6EFCE"),   # 연두
    ("Blocked", "F4CCCC"),   # 연빨강
]
ISSUE_CF_RULES = [
    ("Open",   "FFC7CE"),    # 연빨강
    ("Closed", "C6EFCE"),    # 연두
]


# ══════════════════════════════════════════════════════════════════════
#  시트별 데이터 정의
# ══════════════════════════════════════════════════════════════════════

SHEET_프로젝트개요 = [
    ["항목",       "내용"],
    ["팀명",       "카데이터"],
    ["팀장",       "최연우"],
    ["팀원",       "권소라, 김지혜, 박회종"],
    ["프로젝트명", "물로간다"],
    ["주제",       "한국 자동차 등록 현황 및 기업 FAQ 조회시스템"],
    ["발표일",     "2026-05-19"],
    ["발표 형식",  "팀당 발표 10분 + Q&A 10분"],
    ["GitHub",     ""],
    ["노션/Slack", ""],
    ["", ""],
    ["── 진행 요약 (작업분담 시트 기준) ──", ""],
    ["전체 작업 수", "=COUNTA(작업분담!A2:A1000)-1"],
    ["Done",         "=COUNTIF(작업분담!G2:G1000,\"Done\")"],
    ["Doing",        "=COUNTIF(작업분담!G2:G1000,\"Doing\")"],
    ["Todo",         "=COUNTIF(작업분담!G2:G1000,\"Todo\")"],
    ["Blocked",      "=COUNTIF(작업분담!G2:G1000,\"Blocked\")"],
    ["완료율 (%)",   "=IFERROR(ROUND(COUNTIF(작업분담!G2:G1000,\"Done\")/MAX(COUNTA(작업분담!A2:A1000)-1,1)*100,1),0)"],
]

SHEET_팀원 = [
    ["이름",   "역할",                         "GitHub", "비고"],
    ["최연우", "팀장 / Streamlit GUI 구현",     "",       ""],
    ["권소라", "MySQL DB 설계 및 데이터 저장",  "",       ""],
    ["김지혜", "크롤링 (bs4, Playwright)",       "",       ""],
    ["박회종", "파이썬-MySQL 연동 조회 시스템", "",       ""],
]

# 시작일 전부 2026-05-14
# 마감일: T-14, T-15만 2026-05-19, 나머지 빈칸
# 상태: 전부 Todo (이제 막 시작)
# 산출물, 비고: 전부 빈칸
SHEET_작업분담 = [
    ["ID",    "영역",     "작업명",                                       "담당자", "시작일",       "마감일",       "상태",  "산출물", "비고"],
    ["T-01", "DB",       "DB 스키마 설계 (ERD 작성)",                    "권소라", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-02", "DB",       "MySQL 테이블 생성 및 검증",                    "권소라", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-03", "DB",       "SQLAlchemy 연동 (db.py)",                      "권소라", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-04", "Crawler",  "국토부 엑셀 크롤러 (MolitCarCrawler)",         "김지혜", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-05", "Crawler",  "수소허브 등록현황 크롤러 (h2hub.or.kr)",       "김지혜", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-06", "Crawler",  "무공해차 FAQ 크롤러 (ev.or.kr)",               "김지혜", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-07", "Crawler",  "현대차 수소차 FAQ 크롤러 (hyundai.com)",       "김지혜", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-08", "Backend",  "크롤러-DB 저장 연동 (CrawlService)",          "박회종", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-09", "Backend",  "APScheduler 자동 크롤링 스케줄 등록",         "박회종", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-10", "Backend",  "조회 메서드 (find_all, find_by_region 등)",    "박회종", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-11", "UI",       "Streamlit 탭 구성 (등록현황/FAQ)",             "최연우", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-12", "UI",       "지역별 등록현황 막대/선 그래프",               "최연우", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-13", "UI",       "FAQ 키워드 검색 기능",                         "최연우", "2026-05-14", "",             "Todo",  "",       ""],
    ["T-14", "Docs",     "발표 슬라이드 작성",                           "전원",   "2026-05-14", "2026-05-19",   "Todo",  "",       ""],
    ["T-15", "Docs",     "README 최종 정리",                             "최연우", "2026-05-14", "2026-05-19",   "Todo",  "",       ""],
]

SHEET_데이터출처 = [
    ["번호", "도메인",          "사이트명",                  "URL",                                                                   "비고"],
    ["1",    "자동차 등록현황", "국토교통부 통계누리",       "https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58",         ""],
    ["2",    "수소차 등록현황", "수소모빌리티 허브 (H2HUB)", "https://h2hub.or.kr/main/stat/stat_use_hCar_apply.do",                  ""],
    ["3",    "수소차 FAQ",      "무공해차 통합누리집",       "https://ev.or.kr/nportal/partcptn/initFaqAction.do",                    ""],
    ["4",    "수소차 FAQ",      "현대차 수소차 구매 FAQ",    "https://www.hyundai.com/kr/ko/e/customer/guide/purchase/estimate-fcev", ""],
]

SHEET_회의록 = [
    ["일자", "참석자", "안건", "결정사항", "다음 액션", "담당"],
    ["",     "",       "",     "",          "",          ""],
    ["",     "",       "",     "",          "",          ""],
]

# 빈 이슈로그 (헤더만)
SHEET_이슈로그 = [
    ["번호", "일자", "내용", "담당", "상태", "해결 방법"],
    ["",     "",     "",     "",     "",     ""],
    ["",     "",     "",     "",     "",     ""],
    ["",     "",     "",     "",     "",     ""],
]


# 시트 정의: (시트명, 데이터, 헤더 배경색 RRGGBB)
WORKSHEETS: list[tuple[str, list[list], str]] = [
    ("프로젝트개요", SHEET_프로젝트개요, "4A90D9"),
    ("팀원",         SHEET_팀원,         "27AE60"),
    ("작업분담",     SHEET_작업분담,     "E67E22"),
    ("데이터출처",   SHEET_데이터출처,   "8E44AD"),
    ("회의록",       SHEET_회의록,       "16A085"),
    ("이슈로그",     SHEET_이슈로그,     "C0392B"),
]


# ══════════════════════════════════════════════════════════════════════
#  드롭다운 유효성 검사
# ══════════════════════════════════════════════════════════════════════

def _make_dv(formula: str, prompt: str) -> DataValidation:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    dv.prompt = prompt
    dv.promptTitle = "선택"
    return dv


def _add_dropdowns(ws, sheet_name: str, num_data_rows: int) -> None:
    last = num_data_rows + 1

    if sheet_name == "작업분담":
        dv_status = _make_dv(TASK_STATUS_LIST,  "Todo / Doing / Done / Blocked")
        dv_status.sqref = f"G2:G{last}"
        ws.add_data_validation(dv_status)

        dv_member = _make_dv(MEMBER_LIST, "팀원 선택")
        dv_member.sqref = f"D2:D{last}"
        ws.add_data_validation(dv_member)

        dv_area = _make_dv(AREA_LIST, "영역 선택")
        dv_area.sqref = f"B2:B{last}"
        ws.add_data_validation(dv_area)

    elif sheet_name == "이슈로그":
        dv_status = _make_dv(ISSUE_STATUS_LIST, "Open / Closed")
        dv_status.sqref = f"E2:E{last}"
        ws.add_data_validation(dv_status)

        dv_member = _make_dv(MEMBER_LIST, "팀원 선택")
        dv_member.sqref = f"D2:D{last}"
        ws.add_data_validation(dv_member)

    elif sheet_name == "회의록":
        dv_member = _make_dv(MEMBER_LIST, "팀원 선택")
        dv_member.sqref = f"F2:F{last}"
        ws.add_data_validation(dv_member)


# ══════════════════════════════════════════════════════════════════════
#  조건부 서식 — 드롭다운 값 바뀌면 색상 자동 변경
# ══════════════════════════════════════════════════════════════════════

def _add_conditional_formatting(ws, sheet_name: str, num_data_rows: int) -> None:
    last = num_data_rows + 1

    def add_rules(col_range: str, rules: list[tuple[str, str]]) -> None:
        for status, bg_color in rules:
            ws.conditional_formatting.add(
                col_range,
                CellIsRule(
                    operator="equal",
                    formula=[f'"{status}"'],
                    fill=PatternFill("solid", fgColor=bg_color),
                ),
            )

    if sheet_name == "작업분담":
        add_rules(f"G2:G{last}", TASK_CF_RULES)
    elif sheet_name == "이슈로그":
        add_rules(f"E2:E{last}", ISSUE_CF_RULES)


# ══════════════════════════════════════════════════════════════════════
#  스타일 헬퍼
# ══════════════════════════════════════════════════════════════════════

# 초기 데이터 표시용 색상 (CF 규칙과 동일하게 유지)
_STATUS_FILL: dict[str, str] = {
    "Todo":    "EBF5FB",
    "Doing":   "FFEB9C",
    "Done":    "C6EFCE",
    "Blocked": "F4CCCC",
    "Open":    "FFC7CE",
    "Closed":  "C6EFCE",
}


def _thin_border() -> Border:
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)


def _apply_sheet(ws, data: list[list], header_hex: str) -> None:
    header_fill = PatternFill("solid", fgColor=header_hex)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    body_font   = Font(size=10)
    center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    border      = _thin_border()

    for row_idx, row_data in enumerate(data, start=1):
        is_header = row_idx == 1
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if is_header:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = center
            else:
                cell.font      = body_font
                cell.alignment = left
                # 초기값이 상태 값이면 색상 미리 적용 (CF가 없는 환경 대비)
                if str(value).strip() in _STATUS_FILL:
                    cell.fill = PatternFill("solid", fgColor=_STATUS_FILL[str(value).strip()])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx in range(1, len(data[0]) + 1):
        max_len = 0
        for row in data:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        width = min(max(max_len * 1.4, 8), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20


# ══════════════════════════════════════════════════════════════════════
#  메인
# ══════════════════════════════════════════════════════════════════════

def create_workbook() -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_name, data, color in WORKSHEETS:
        ws = wb.create_sheet(title=sheet_name)
        _apply_sheet(ws, data, color)
        _add_dropdowns(ws, sheet_name, len(data) - 1)
        _add_conditional_formatting(ws, sheet_name, len(data) - 1)
        print(f"  ✔ {sheet_name}")

    wb.save(OUTPUT_PATH)
    return OUTPUT_PATH


def main() -> int:
    print("── 물로간다 팀작업현황 Excel 생성 ──")
    path = create_workbook()
    print(f"\n저장 완료: {path.resolve()}")
    print()
    print("▶ 구글 스프레드시트로 올리는 방법")
    print("  1. https://drive.google.com → 새로 만들기 → 파일 업로드")
    print(f"     → {path.name}  선택")
    print("  2. 업로드된 파일 우클릭 → 'Google 스프레드시트로 열기'")
    print("  3. 공유 버튼 → '링크가 있는 모든 사용자' → '편집자'")
    print("     → 링크 복사 → 팀 단톡방에 공유")
    return 0


if __name__ == "__main__":
    sys.exit(main())
