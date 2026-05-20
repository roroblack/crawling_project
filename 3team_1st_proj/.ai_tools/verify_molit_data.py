"""
국토부 수소차 데이터 검증 스크립트

1. MolitCarCrawler 를 이용해 연도별 12월 파일을 내려받는다.
2. 내려받은 xlsx 파일을 molit_downloads/ 에 <연도>년_<월>월_자동차등록.xlsx 로 저장한다.
3. 각 파일에서 수소/수소전기 소계 행을 추출한다.
4. 결과를 molit_downloads/검증결과.md 로 저장한다.

실행:
  .venv\\Scripts\\python.exe verify_molit_data.py
"""

import os
import io
import asyncio
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor

import openpyxl
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
from dotenv import load_dotenv

load_dotenv()

# ── 설정 ──────────────────────────────────────────────────────
DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "molit_downloads")
URL          = "https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58"
# 확인할 연도 목록 (12월) + 현재 월
YEARS_BACK   = 8     # 현재 기준 몇 년치까지
# ──────────────────────────────────────────────────────────────

os.makedirs(DOWNLOAD_DIR, exist_ok=True)


def _build_year_list() -> list[tuple[int, int]]:
    """현재 월 + 과거 YEARS_BACK년치 12월 목록"""
    now = datetime.now()
    result = [(now.year, now.month)]
    for i in range(1, YEARS_BACK + 1):
        result.append((now.year - i, 12))
    return result


def _run_playwright_html() -> str:
    loop = asyncio.ProactorEventLoop()
    return loop.run_until_complete(_async_get_html())


async def _async_get_html() -> str:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        page = await browser.new_page(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
            )
        )
        await page.goto(URL, wait_until="networkidle", timeout=60_000)
        await page.wait_for_timeout(1_500)
        html = await page.content()
        await browser.close()
    return html


def _parse_links(html: str) -> list[dict]:
    soup = BeautifulSoup(html, "lxml")
    links = []
    for a in soup.select("a[onclick*='downFile']"):
        links.append({
            "text":    a.get_text(strip=True),
            "onclick": a.get("onclick", ""),
        })
    return links


def _find_link(links: list[dict], year: int, month: int) -> dict | None:
    patterns = [
        f"{year}년 {month}월 자동차 등록자료",
        f"{year}년 {month:02d}월 자동차 등록자료",
    ]
    for link in links:
        text = link.get("text", "")
        if any(p in text for p in patterns):
            return link
    return None


def _run_download(onclick: str) -> bytes | None:
    loop = asyncio.ProactorEventLoop()
    return loop.run_until_complete(_async_download(onclick))


async def _async_download(onclick: str) -> bytes | None:
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx  = await browser.new_context(accept_downloads=True)
        page = await ctx.new_page()
        await page.goto(URL, wait_until="networkidle", timeout=60_000)
        try:
            async with page.expect_download(timeout=60_000) as dl_info:
                await page.eval_on_selector_all(
                    "a[onclick*='downFile']",
                    "(els, oc) => { for (const e of els) { if (e.getAttribute('onclick') === oc) { e.click(); break; } } }",
                    onclick,
                )
            download = await dl_info.value
            path = await download.path()
            with open(path, "rb") as f:
                data = f.read()
        except Exception as e:
            print(f"  [WARN] 다운로드 실패: {e}")
            data = None
        finally:
            await browser.close()
    return data


HYDROGEN_FUELS = {"수소", "수소전기"}
REGIONS_COL = [
    "서울", "부산", "대구", "인천", "광주", "대전", "울산", "세종",
    "경기", "강원", "충북", "충남", "전북", "전남", "경북", "경남", "제주",
]


def _parse_hydrogen(xlsx_bytes: bytes, year: int, month: int) -> dict:
    """
    10.연료별_등록현황 시트에서 수소/수소전기 '소계 | 계' 행만 추출한다.

    2023년 이후 신규 형식:
      A=수소  B=소계  C=계  D..=지역별 값  (한 행에 모두 명시)
    2022년 이전 구버전 형식:
      A=수소  B=None  C=None  (병합 헤더)
      ...
      A=None  B=소계  C=비사업용  (소계 섹션 첫 행)
      A=None  B=None  C=계        ← 이 행이 실제 소계 전체 합계 행

    두 형식 모두 current_fuel / current_vehicle 컨텍스트를 추적해 처리한다.

    반환: {region: count, ..., '전국': count}
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)

    # 연료별 시트명은 연도에 따라 공백 포함 가능("10.연료별_등 록현황")
    fuel_sheet = None
    for name in wb.sheetnames:
        if "연료" in name:
            fuel_sheet = name
            break
    if not fuel_sheet:
        return {}

    ws   = wb[fuel_sheet]
    rows = list(ws.iter_rows(values_only=True))

    # 헤더 행 탐색 (서울·경기 둘 다 포함된 행)
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip() if c else "" for c in row]
        if "서울" in cells and "경기" in cells:
            header_idx = i
            break
    if header_idx is None:
        return {}

    header = [str(c).strip() if c else "" for c in rows[header_idx]]

    # 지역명 → 열 인덱스 (계 = 전국 합계 열)
    col_map: dict[str, int] = {}
    for j, h in enumerate(header):
        if h in REGIONS_COL:
            col_map[h] = j
        if h == "계" and "전국" not in col_map and col_map:
            col_map["전국"] = j

    # 컨텍스트 추적 파싱
    total: dict[str, int] = {r: 0 for r in list(col_map.keys())}
    current_fuel    = None
    current_vehicle = None

    for row in rows[header_idx + 1:]:
        if not row:
            continue

        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        c = str(row[2]).strip() if row[2] else ""

        # A열이 비어있지 않으면 새 연료 섹션 시작
        if a:
            current_fuel    = a
            current_vehicle = None
        # B열이 비어있지 않으면 차종 컨텍스트 갱신
        if b:
            current_vehicle = b

        # 수소/수소전기 섹션의 소계/계 행인지 확인
        # - 신규 형식: a=수소, b=소계, c=계 → is_target
        # - 구버전:   current_fuel=수소, current_vehicle=소계, c=계 → is_target
        if current_fuel not in HYDROGEN_FUELS:
            continue
        if current_vehicle != "소계":
            continue
        if c != "계":
            continue

        for region, col_idx in col_map.items():
            if col_idx < len(row):
                raw = row[col_idx]
                total[region] += int(raw) if isinstance(raw, (int, float)) and raw == raw else 0

    return total


def main():
    print("=" * 60)
    print("국토부 수소차 등록현황 검증 스크립트")
    print(f"실행 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. 페이지 HTML 가져오기 (1회)
    print("\n[1/3] 통계누리 페이지 로딩 중...")
    with ThreadPoolExecutor(max_workers=1) as pool:
        html = pool.submit(_run_playwright_html).result()
    links = _parse_links(html)
    print(f"      다운로드 링크 {len(links)}개 발견")

    year_list = _build_year_list()
    print(f"\n[2/3] 수집 대상: {year_list}")

    results: list[dict] = []   # {year, month, regions:{...}, save_path}

    for year, month in year_list:
        label = f"{year}년 {month:02d}월"

        # 이미 내려받은 파일이 있으면 재사용
        save_path = os.path.join(DOWNLOAD_DIR, f"{year}년_{month:02d}월_자동차등록.xlsx")
        if os.path.exists(save_path):
            print(f"  [{label}] 기존 파일 사용: {os.path.basename(save_path)}")
            with open(save_path, "rb") as f:
                xlsx_bytes = f.read()
        else:
            link = _find_link(links, year, month)
            if link is None:
                print(f"  [{label}] 링크 없음 — 건너뜀")
                continue
            print(f"  [{label}] 다운로드 중...", end=" ", flush=True)
            with ThreadPoolExecutor(max_workers=1) as pool:
                xlsx_bytes = pool.submit(_run_download, link["onclick"]).result()
            if xlsx_bytes is None:
                print("실패")
                continue
            with open(save_path, "wb") as f:
                f.write(xlsx_bytes)
            print(f"저장 완료 ({len(xlsx_bytes)//1024} KB)")

        region_counts = _parse_hydrogen(xlsx_bytes, year, month)
        results.append({
            "year":      year,
            "month":     month,
            "label":     label,
            "regions":   region_counts,
            "save_path": save_path,
        })

    # 3. 검증 보고서 작성
    print("\n[3/3] 검증 보고서 작성 중...")
    report_path = os.path.join(DOWNLOAD_DIR, "검증결과.md")
    _write_report(results, report_path)
    print(f"      보고서 저장: {report_path}")
    print("\n완료.")


def _write_report(results: list[dict], path: str) -> None:
    region_order = ["전국"] + REGIONS_COL

    lines = [
        "# 국토부 수소차 등록현황 검증 결과",
        "",
        f"- 생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"- 출처: https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58",
        f"- 통계명: 자동차등록현황보고 (Total Registered Motor Vehicles)",
        f"- 시트: 10.연료별_등록현황 → 수소+수소전기 소계|계 합산",
        "",
        "---",
        "",
        "## 연도별 수소차 등록 대수 (전국 + 17개 시도)",
        "",
    ]

    # 전국 합계 테이블
    lines.append("### 전국 합계 추이")
    lines.append("")
    lines.append("| 연도/월 | 전국 합계 | 전년 대비 증감 | 파일명 |")
    lines.append("|---|---|---|---|")
    prev_national = None
    for r in sorted(results, key=lambda x: (x["year"], x["month"])):
        national = r["regions"].get("전국", 0)
        diff_str = ""
        if prev_national is not None:
            diff = national - prev_national
            diff_str = f"+{diff:,}" if diff >= 0 else f"{diff:,}"
        lines.append(
            f"| {r['label']} | {national:,} | {diff_str} | {os.path.basename(r['save_path'])} |"
        )
        prev_national = national

    lines += ["", "---", "", "## 지역별 상세 (수소+수소전기 합산)", ""]

    # 지역별 상세 테이블
    header_row = "| 지역 | " + " | ".join(
        r["label"] for r in sorted(results, key=lambda x: (x["year"], x["month"]))
    ) + " |"
    sep_row    = "|---|" + "---|" * len(results)
    lines.append(header_row)
    lines.append(sep_row)

    sorted_results = sorted(results, key=lambda x: (x["year"], x["month"]))
    for region in region_order:
        row_vals = [f"{r['regions'].get(region, 0):,}" for r in sorted_results]
        lines.append(f"| {region} | " + " | ".join(row_vals) + " |")

    lines += [
        "",
        "---",
        "",
        "## 파싱 방식 설명",
        "",
        "- 각 xlsx 파일의 `10.연료별_등록현황` 시트에서 헤더 행(서울·경기 모두 포함된 행)을 탐색",
        "- 연료 = `수소` 또는 `수소전기`, 차종 = `소계`, 용도 = `계` 인 행만 추출",
        "- 두 연료의 값을 **합산**하여 지역별 수소차 총 보유 대수를 산출",
        "- 각 월 파일의 수치는 **해당 월 말일 기준 누적 등록 대수 스냅샷**",
        "  (출처 통계 메타: 작성대상월 말일 기준)",
        "- 연도별 추이: 매년 12월 스냅샷을 비교하면 연간 보급 성장세 확인 가능",
        "",
    ]

    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


if __name__ == "__main__":
    main()
