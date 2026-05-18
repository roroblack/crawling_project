"""
crawler_molit.py _parse() 알고리즘 검증 스크립트
소계/계 직접 읽기 방식과 DB 저장 데이터를 비교한다.
"""
import sys, os, glob, re
sys.path.insert(0, '.')

import openpyxl, io
from crawler_molit import MolitCarCrawler

c = MolitCarCrawler()
files = sorted(f for f in glob.glob('molit_downloads/*.xlsx') if not os.path.basename(f).startswith('~$'))

results = {}
for fp in files:
    m = re.search(r'(\d{4})', os.path.basename(fp))
    if not m:
        continue
    yr = int(m.group(1))
    with open(fp, 'rb') as fh:
        data = fh.read()
    items = c._parse(data, yr, c.TARGET_FUELS)
    r = {it.region: it.count for it in items}
    results[yr] = r

# 전체 연도 출력
print("=== 연도별 전체 검증 ===")
header = f"{'연도':>4}  {'전국':>7}  {'서울':>6}  {'부산':>6}  {'경기':>7}  {'제주':>5}"
print(header)
print("-" * len(header))
for yr in sorted(results):
    r = results[yr]
    total  = r.get('전국', 0)
    seoul  = r.get('서울', 0)
    busan  = r.get('부산', 0)
    gg     = r.get('경기', 0)
    jeju   = r.get('제주', 0)
    print(f"{yr:>4}  {total:>7,}  {seoul:>6,}  {busan:>6,}  {gg:>7,}  {jeju:>5,}")

# 2025 전체 지역 출력
print()
print("=== 2025 지역별 상세 ===")
if 2025 in results:
    for region, cnt in sorted(results[2025].items(), key=lambda x: -x[1]):
        print(f"  {region}: {cnt:,}")

# 2025 소계/계 행 원본값 직접 출력 (수소/수소전기)
print()
print("=== 2025 소계/계 행 원본값 (수소/수소전기 모두) ===")
yr = 2025
fp_list = [f for f in files if str(yr) in os.path.basename(f)]
if fp_list:
    wb = openpyxl.load_workbook(fp_list[0], read_only=True, data_only=True)
    ws = wb['10.연료별_등록현황']
    rows = list(ws.iter_rows(values_only=True))
    current_fuel = ""
    current_vehicle = ""
    for i, row in enumerate(rows):
        a = str(row[0]).strip() if row[0] else ""
        b = str(row[1]).strip() if row[1] else ""
        c_val = str(row[2]).strip() if row[2] else ""
        if a:
            current_fuel = a
            current_vehicle = ""
        if b:
            current_vehicle = b
        if current_fuel in ('수소', '수소전기') and current_vehicle == '소계' and c_val == '계':
            vals = list(row[3:22])
            print(f"  row{i}: {current_fuel}/소계/계 -> {vals}")
    wb.close()

# 2018 구조 확인 (병합 셀 forward-fill 검증)
print()
print("=== 2018 수소 섹션 구조 (병합 셀 확인) ===")
yr = 2018
fp_list = [f for f in files if str(yr) in os.path.basename(f)]
if fp_list:
    wb = openpyxl.load_workbook(fp_list[0], read_only=True, data_only=True)
    ws = wb['10.연료별_등록현황']
    rows = list(ws.iter_rows(values_only=True))
    in_h = False
    for i, row in enumerate(rows):
        a = str(row[0]).strip() if row[0] else ''
        b = str(row[1]).strip() if row[1] else ''
        c_val = str(row[2]).strip() if row[2] else ''
        if a == '수소':
            in_h = True
        if in_h:
            vals = list(row[3:21])
            print(f"  row{i:3d}: A={a!r:10} B={b!r:8} C={c_val!r:8} | {vals}")
        if in_h and a and a != '수소':
            break
    wb.close()
