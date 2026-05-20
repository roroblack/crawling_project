import asyncio
import io
import os
from playwright.async_api import async_playwright


DOWNLOAD_DIR = os.path.join(os.path.dirname(__file__), "molit_downloads")
os.makedirs(DOWNLOAD_DIR, exist_ok=True)


async def main():
    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        ctx = await browser.new_context(accept_downloads=True)
        page = await ctx.new_page()

        print("페이지 로딩 중...")
        await page.goto("https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58", timeout=60000)
        await page.wait_for_load_state("networkidle", timeout=30000)

        # downFile 링크 전부 찾기
        links = await page.eval_on_selector_all(
            "a[onclick*='downFile']",
            "els => els.map(e => ({ text: e.textContent.trim(), onclick: e.getAttribute('onclick') }))"
        )
        print(f"\n총 다운로드 링크: {len(links)}개")
        
        # 자동차 등록자료 첫 번째 파일만 다운로드
        car_links = [l for l in links if "자동차 등록자료" in l["text"]]
        print(f"자동차 등록자료 링크: {len(car_links)}개")
        if car_links:
            print(f"\n최신 3개:")
            for l in car_links[:3]:
                print(f"  {l['text']}")

        if not car_links:
            print("자동차 등록자료 링크 없음")
            await browser.close()
            return

        # 가장 최신 파일 다운로드
        target_text = car_links[0]["text"]
        print(f"\n다운로드 시도: {target_text}")

        # download 이벤트 대기
        async with page.expect_download(timeout=30000) as dl_info:
            await page.eval_on_selector_all(
                "a[onclick*='downFile']",
                f"els => {{ for(let e of els) {{ if(e.textContent.trim().includes('자동차 등록자료')) {{ e.click(); break; }} }} }}"
            )
        download = await dl_info.value
        save_path = os.path.join(DOWNLOAD_DIR, download.suggested_filename)
        await download.save_as(save_path)
        print(f"저장 완료: {save_path}")
        file_size = os.path.getsize(save_path)
        print(f"파일 크기: {file_size:,} bytes")

        # openpyxl로 시트 구조 확인
        import openpyxl
        wb = openpyxl.load_workbook(save_path, read_only=True, data_only=True)
        print(f"\n===== 시트 목록 ({len(wb.sheetnames)}개) =====")
        print(wb.sheetnames)

        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(max_row=5, values_only=True))
            print(f"\n{'='*60}")
            print(f"시트: {sn}")
            print(f"{'='*60}")
            for row in rows:
                vals = [str(c) if c is not None else "" for c in row[:16]]
                # 빈 행 스킵
                if any(v.strip() for v in vals):
                    print(" | ".join(vals))

        await browser.close()


asyncio.run(main())
