# 운영체제 환경변수 값을 읽기 위해 os 모듈을 가져온다.
import os
import io

# 현재 날짜와 시간을 저장하기 위해 datetime을 가져온다.
from datetime import datetime

# 엑셀 파일을 파싱하기 위해 openpyxl을 가져온다.
import openpyxl

# HTML 문서를 분석하기 위해 BeautifulSoup을 가져온다.
from bs4 import BeautifulSoup

# .env 파일에 저장된 환경변수를 읽기 위해 load_dotenv를 가져온다.
from dotenv import load_dotenv

# Playwright 비동기 API를 사용한다.
from playwright.async_api import async_playwright

# Streamlit 환경에서 asyncio 이벤트 루프를 별도 스레드에서 실행하기 위해 가져온다.
from concurrent.futures import ThreadPoolExecutor

# models.py에 정의된 CarRegistrationItem 데이터 클래스를 가져온다.
from models import CarRegistrationItem

from db import REGIONS


# .env 파일을 읽어서 os.getenv()로 사용할 수 있게 한다.
load_dotenv()


# 국토교통부 통계누리(stat.molit.go.kr)에서 자동차 등록 현황 엑셀을 내려받아
# 수소차 데이터를 추출하는 크롤러이다.
class MolitCarCrawler:
    # 10.연료별_등록현황 시트의 지역 열 이름 목록 (헤더 행 기준)
    REGION_COLS = REGIONS

    # 수집할 연료 종류 기본값 (수소차 관련만)
    TARGET_FUELS = {"수소", "수소전기"}

    # 객체가 생성될 때 자동으로 실행되는 생성자이다.
    def __init__(self):
        # 크롤링 대상 페이지 URL (자동차 등록자료 통계, hRsId=58)
        self.url = "https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58"

        # 엑셀 임시 저장 폴더 (프로젝트 내부, .gitignore에 추가 권장)
        self.download_dir = os.path.join(
            os.path.dirname(__file__), "..", "molit_downloads"
        )
        os.makedirs(self.download_dir, exist_ok=True)

    # 수소차 연도별 등록현황을 수집하는 대표 메서드이다.
    # 당월 등록된 자동차 수가 전체 누계니까 연도별 12월 데이터만 수집한다.
    # years_back년치 연말(12월) 파일 + 현재 월 파일만 내려받아
    # 10.연료별_등록현황 시트에서 수소/수소전기 행을 추출한다.
    # 전체 월 수집(96개) 대신 연말 스냅샷(9개)만 받으므로 속도가 빠르다.
    # fuel_filter: None이면 TARGET_FUELS(수소류)만, 빈 set이면 전체 연료 수집
    # crawl() 실행 후 실제로 다운로드된 현재 연도의 데이터 월 (예: 4 = 4월)
    # 현재 연도 데이터를 찾지 못했으면 None
    last_stat_month: int | None = None

    def crawl(
        self,
        years_back: int = 8,
        fuel_filter: set | None = None,
    ) -> list[CarRegistrationItem]:

        if fuel_filter is None:
            fuel_filter = self.TARGET_FUELS

        # 페이지 HTML을 가져온 뒤 링크 목록을 한 번만 추출한다.
        html    = self._get_html()
        links   = self._parse_links(html)

        all_items: list[CarRegistrationItem] = []
        self.last_stat_month = None
        now = datetime.now()

        # ── 현재 연도: 최신 월부터 1월까지 역방향으로 탐색 ─────────────
        for m in range(now.month, 0, -1):
            link = self._find_link(links, now.year, m)
            if link is None:
                continue
            xlsx_bytes = self._get_file(link["onclick"])
            if xlsx_bytes is None:
                continue
            items = self._parse(xlsx_bytes, now.year, fuel_filter)
            if items:
                all_items.extend(items)
                self.last_stat_month = m
            break

        # ── 과거 연도: 12월(연말 스냅샷)만 수집 ─────────────────────────
        for i in range(1, years_back + 1):
            year  = now.year - i
            month = 12
            link = self._find_link(links, year, month)
            if link is None:
                continue
            xlsx_bytes = self._get_file(link["onclick"])
            if xlsx_bytes is None:
                continue
            items = self._parse(xlsx_bytes, year, fuel_filter)
            all_items.extend(items)

        return all_items

    # ── 내부 헬퍼 메서드들 ──────────────────────────────────────────────

    def _build_year_list(self, years_back: int) -> list[tuple[int, int]]:
        now    = datetime.now()
        result = []
        for i in range(1, years_back + 1):
            result.append((now.year - i, 12))
        return result

    def _get_html(self) -> str:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._run_playwright)
            return future.result()

    def _run_playwright(self) -> str:
        import asyncio
        loop = asyncio.ProactorEventLoop()
        return loop.run_until_complete(self._async_get_html())

    async def _async_get_html(self) -> str:
        async with async_playwright() as p:
            browser     = await p.chromium.launch(headless=True)
            page        = await browser.new_page(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
                )
            )
            await page.goto(self.url, wait_until="networkidle", timeout=60000)
            await page.wait_for_timeout(1500)
            html = await page.content()
            await browser.close()
        return html

    def _parse_links(self, html: str) -> list[dict]:
        soup    = BeautifulSoup(html, "lxml")
        links   = []
        for a in soup.select("a[onclick*='downFile']"):
            links.append({
                "text": a.get_text(strip=True),
                "onclick": a.get("onclick", ""),
            })
        return links

    def _find_link(self, links: list[dict], year: int, month: int) -> dict | None:
        patterns = [
            f"{year}년 {month}월 자동차 등록자료",
            f"{year}년 {month:02d}월 자동차 등록자료",
        ]
        for link in links:
            text = link.get("text", "")
            if any(p in text for p in patterns):
                return link
        return None

    def _get_file(self, onclick: str) -> bytes | None:
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._run_download, onclick)
            return future.result()

    def _run_download(self, onclick: str) -> bytes | None:
        import asyncio
        loop = asyncio.ProactorEventLoop()
        return loop.run_until_complete(self._async_download(onclick))

    async def _async_download(self, onclick: str) -> bytes | None:
        async with async_playwright() as p:
            browser     = await p.chromium.launch(headless=True)
            ctx         = await browser.new_context(accept_downloads=True)
            page        = await ctx.new_page()
            await page.goto(self.url, wait_until="networkidle", timeout=60000)

            try:
                async with page.expect_download(timeout=30000) as dl_info:
                    await page.eval_on_selector_all(
                        "a[onclick*='downFile']",
                        "(els, oc) => { for (const e of els) { if (e.getAttribute('onclick') === oc) { e.click(); break; } } }",
                        onclick,
                    )
                download = await dl_info.value
                path = await download.path()
                with open(path, "rb") as f:
                    data = f.read()
            except Exception:
                data = None
            finally:
                await browser.close()

        return data

    def _parse(
        self,
        xlsx_bytes: bytes,
        year: int,
        fuel_filter: set,
    ) -> list[CarRegistrationItem]:
        # ──────────────────────────────────────────────────────────────
        # [LEGACY v1] 개별 행 합산 방식
        # 비사업용 + 사업용 각 행을 직접 읽어 차종별로 합산한다.
        # 소계(B='소계') · 합계(C='계') 행은 중복이므로 제외한다.
        #
        # 문제점: 2018-2022 Excel은 B열이 병합 셀이어서 하위 행의 B=None.
        #         forward-fill(current_vehicle)을 추가하면 정상 동작한다.
        # 현재(v2) 방식: 소계/계 행 직접 읽기 → crawler_molit.py 참조
        # ──────────────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
        if "10.연료별_등록현황" not in wb.sheetnames:
            return []

        ws = wb["10.연료별_등록현황"]
        all_rows = list(ws.iter_rows(values_only=True))

        # 헤더 행(지역명이 있는 행)을 찾는다.
        header_row_idx = None
        for i, row in enumerate(all_rows):
            cells = [str(c).strip() if c else "" for c in row]
            if "서울" in cells and "경기" in cells:
                header_row_idx = i
                break

        if header_row_idx is None:
            return []

        # 헤더 행에서 지역명 → 열 인덱스 매핑을 만든다.
        header = all_rows[header_row_idx]
        region_col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(header):
            cell_str = str(cell).strip() if cell else ""
            if cell_str in self.REGION_COLS:
                region_col_map[cell_str] = col_idx
            if cell_str == "계" and "전국" not in region_col_map and region_col_map:
                region_col_map["전국"] = col_idx

        # 지역별 합산 딕셔너리
        region_totals: dict[str, int] = {r: 0 for r in region_col_map}

        # 데이터 행을 순회하며 지정된 연료 행의 값을 지역별로 합산한다.
        # ─ 병합 셀 대응: A열(연료) · B열(차종)이 None이면 이전 값을 유지한다.
        #   2018-2022 파일은 같은 연료/차종 그룹의 행을 병합 셀로 표현하므로
        #   첫 행을 제외한 나머지 행은 A·B가 None 으로 읽힌다.
        current_fuel    = ""
        current_vehicle = ""

        for row in all_rows[header_row_idx + 1:]:
            if not row:
                continue

            # A열: 연료 — 새 값이 있으면 갱신 및 차종 초기화
            fuel_raw = str(row[0]).strip() if row[0] else ""
            if fuel_raw:
                current_fuel    = fuel_raw
                current_vehicle = ""           # 연료 섹션이 바뀌면 차종 초기화
            fuel = current_fuel

            if not fuel:
                continue                       # 아직 연료가 확정되지 않은 초기 행

            if fuel_filter and fuel not in fuel_filter:
                continue                       # 대상 외 연료는 건너뜀

            # B열: 차종 — 새 값이 있으면 갱신 (병합 셀이면 이전 차종 유지)
            vehicle_raw = str(row[1]).strip() if row[1] else ""
            if vehicle_raw:
                current_vehicle = vehicle_raw
            vehicle = current_vehicle

            usage = str(row[2]).strip() if row[2] else ""     # C열: 용도

            # 소계(B='소계') · 합계(C='계') 행은 중복 집계이므로 제외
            if not vehicle or not usage or vehicle == "소계" or usage == "계":
                continue

            for region, col_idx in region_col_map.items():
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                count = int(raw) if isinstance(raw, (int, float)) and raw == raw else 0
                region_totals[region] = region_totals.get(region, 0) + count

        return [
            CarRegistrationItem(stat_year=year, region=r, count=c)
            for r, c in region_totals.items()
            if c > 0
        ]
