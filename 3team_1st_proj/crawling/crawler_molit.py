# 운영체제 환경변수 값을 읽기 위해 os 모듈을 가져온다.
import os
import io

# 현재 날짜와 시간을 저장하기 위해 datetime을 가져온다.
from datetime import datetime

# 엑셀 파일을 파싱하기 위해 openpyxl을 가져온다.
import openpyxl

# HTML 문서를 분석하기 위해 BeautifulSoup을 가져온다.
from bs4 import BeautifulSoup

# .env 파일에 저장된 환경변수를 읽기 위해 load_dotenv를 가져온다.
from dotenv import load_dotenv

# Playwright 비동기 API를 사용한다.
from playwright.async_api import async_playwright

# Streamlit 환경에서 asyncio 이벤트 루프를 별도 스레드에서 실행하기 위해 가져온다.
from concurrent.futures import ThreadPoolExecutor

# models.py에 정의된 CarRegistrationItem 데이터 클래스를 가져온다.
from model.models import CarRegistrationItem

from common.db import REGIONS


# .env 파일을 읽어서 os.getenv()로 사용할 수 있게 한다.
load_dotenv()


# 국토교통부 통계누리(stat.molit.go.kr)에서 자동차 등록 현황 엑셀을 내려받아
# 수소차 데이터를 추출하는 크롤러이다.
class MolitCarCrawler:
    # 10.연료별_등록현황 시트의 지역 열 이름 목록 (헤더 행 기준)
    REGION_COLS = REGIONS

    # 수집할 연료 종류 기본값 (수소차 관련만)
    TARGET_FUELS = {"수소", "수소전기"}

    # 객체가 생성될 때 자동으로 실행되는 생성자이다.
    def __init__(self):
        # 크롤링 대상 페이지 URL (자동차 등록자료 통계, hRsId=58)
        self.url = "https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58"

        # 엑셀 임시 저장 폴더 (프로젝트 내부, .gitignore에 추가 권장)
        self.download_dir = os.path.join(
            os.path.dirname(__file__), "molit_downloads"
        )
        os.makedirs(self.download_dir, exist_ok=True)

    # 수소차 연도별 등록현황을 수집하는 대표 메서드이다.
    # 당월 등록된 자동차 수가 전체 누계니까 연도별 12월 데이터만 수집한다. 
    # years_back년치 연말(12월) 파일 + 현재 월 파일만 내려받아 
    # 10.연료별_등록현황 시트에서 수소/수소전기 행을 추출한다.
    # 전체 월 수집(96개) 대신 연말 스냅샷(9개)만 받으므로 속도가 빠르다.
    # fuel_filter: None이면 TARGET_FUELS(수소류)만, 빈 set이면 전체 연료 수집
    # crawl() 실행 후 실제로 다운로드된 현재 연도의 데이터 월 (예: 4 = 4월)
    # 현재 연도 데이터를 찾지 못했으면 None
    last_stat_month: int | None = None

    def crawl(
        self,
        years_back: int = 8,
        fuel_filter: set | None = None,
    ) -> list[CarRegistrationItem]:

        if fuel_filter is None:
            fuel_filter = self.TARGET_FUELS

        # 페이지 HTML을 가져온 뒤 링크 목록을 한 번만 추출한다.
        html    = self._get_html()
        links   = self._parse_links(html)

        all_items: list[CarRegistrationItem] = []
        self.last_stat_month = None
        now = datetime.now()

        # ── 현재 연도: 최신 월부터 1월까지 역방향으로 탐색 ─────────────
        # 정부 통계 파일은 당월 공개까지 시차가 있으므로
        # 현재 월 파일이 없으면 전 월 → 전전 월 순으로 시도한다.
        for m in range(now.month, 0, -1):
            link = self._find_link(links, now.year, m)
            if link is None:
                continue
            xlsx_bytes = self._get_file(link["onclick"])
            if xlsx_bytes is None:
                continue
            items = self._parse(xlsx_bytes, now.year, fuel_filter)
            if items:
                all_items.extend(items)
                self.last_stat_month = m   # 실제 데이터 월 기록
            break  # 현재 연도는 하나의 월만 사용

        # ── 과거 연도: 12월(연말 스냅샷)만 수집 ─────────────────────────
        for i in range(1, years_back + 1):
            year  = now.year - i
            month = 12
            link = self._find_link(links, year, month)
            if link is None:
                continue
            xlsx_bytes = self._get_file(link["onclick"])
            if xlsx_bytes is None:
                continue
            items = self._parse(xlsx_bytes, year, fuel_filter)
            all_items.extend(items)

        return all_items

    # ── 내부 헬퍼 메서드들 ──────────────────────────────────────────────

    def _build_year_list(self, years_back: int) -> list[tuple[int, int]]:
        # 과거 years_back년의 12월(연말 스냅샷) 목록만 반환한다.
        # 현재 연도 탐색은 crawl() 에서 역방향 월 탐색으로 직접 처리한다.
        now    = datetime.now()
        result = []
        for i in range(1, years_back + 1):
            result.append((now.year - i, 12))
        return result

    def _get_html(self) -> str:
        # 웹페이지 HTML을 가져온다.
        # Streamlit은 asyncio 이벤트 루프를 사용하기 때문에
        # 별도의 스레드에서 실행하면 NotImplementedError를 피할 수 있다.
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._run_playwright)
            return future.result()

    def _run_playwright(self) -> str:
        # Windows에서 ProactorEventLoop을 명시적으로 생성하여
        # Streamlit/tornado의 SelectorEventLoop 정책과의 충돌을 피한다.
        import asyncio
        loop = asyncio.ProactorEventLoop()
        return loop.run_until_complete(self._async_get_html())

    # async_playwright를 사용해 비동기로 HTML을 가져온다.
    # 기다리는 페이지를 기다리는 동안 Streamlit UI가 멈추지 않게 asyncio 이벤트 루프를 활용한다.
    async def _async_get_html(self) -> str:
        # async_playwright를 사용해 비동기로 HTML을 가져온다.
        async with async_playwright() as p:
            browser     = await p.chromium.launch(headless=True)
            page        = await browser.new_page(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
                )
            )
            await page.goto(self.url, wait_until="networkidle", timeout=60000)
            await page.wait_for_timeout(1500)
            html = await page.content()
            await browser.close()
        return html

    def _parse_links(self, html: str) -> list[dict]:
        # HTML에서 downFile 링크 정보를 추출한다.
        # 각 항목은 {"text": 표시 텍스트, "onclick": onclick 속성값} 형태이다.
        soup    = BeautifulSoup(html, "lxml")
        links   = []
        for a in soup.select("a[onclick*='downFile']"):
            links.append({
                "text": a.get_text(strip=True),
                "onclick": a.get("onclick", ""),
            })
        return links

    def _find_link(self, links: list[dict], year: int, month: int) -> dict | None:
        # 링크 텍스트에서 연도와 월이 일치하는 자동차 등록자료 항목을 찾는다.
        # 파일명 패턴 예시:
        #   "2026년 4월 자동차 등록자료 통계.xlsx"
        #   "2026년 04월 자동차 등록자료 통계.xlsx"  (0 패딩 있는 경우)
        patterns = [
            f"{year}년 {month}월 자동차 등록자료",
            f"{year}년 {month:02d}월 자동차 등록자료",
        ]
        for link in links:
            text = link.get("text", "")
            if any(p in text for p in patterns):
                return link
        return None

    def _get_file(self, onclick: str) -> bytes | None:
        # onclick 속성값을 전달해 Playwright가 해당 링크를 클릭 후 파일을 내려받는다.
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._run_download, onclick)
            return future.result()

    def _run_download(self, onclick: str) -> bytes | None:
        # 다운로드 실행 시에도 별도의 이벤트 루프를 사용해 Streamlit UI가 멈추지 않게 한다.
        import asyncio
        loop = asyncio.ProactorEventLoop()
        return loop.run_until_complete(self._async_download(onclick))

    async def _async_download(self, onclick: str) -> bytes | None:
        async with async_playwright() as p:
            browser     = await p.chromium.launch   (headless=True)
            ctx         = await browser.new_context (accept_downloads=True)
            page        = await ctx.new_page        ()
            await page.goto(self.url, wait_until="networkidle", timeout=60000)

            try:
                # onclick 속성이 동일한 링크를 클릭해 다운로드 이벤트를 유도한다.
                async with page.expect_download(timeout=30000) as dl_info:
                    await page.eval_on_selector_all(
                        "a[onclick*='downFile']",
                        "(els, oc) => { for (const e of els) { if (e.getAttribute('onclick') === oc) { e.click(); break; } } }",
                        onclick,
                    )
                download = await dl_info.value

                # 다운로드된 파일을 bytes로 읽는다 (디스크 저장 없이 메모리로 처리).
                path = await download.path()
                with open(path, "rb") as f:
                    data = f.read()
            except Exception:
                data = None
            finally:
                await browser.close()

        return data

    def _parse(
        self,
        xlsx_bytes: bytes,
        year: int,
        fuel_filter: set,
    ) -> list[CarRegistrationItem]:
        # 10.연료별_등록현황 시트를 파싱해 지역별 수소차 누적 합산 CarRegistrationItem 목록을 반환한다.
        # fuel_type / vehicle_type / usage_type 는 내부 필터링에만 사용하고 출력에는 포함하지 않는다.
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
        if "10.연료별_등록현황" not in wb.sheetnames:
            return []

        ws = wb["10.연료별_등록현황"]
        all_rows = list(ws.iter_rows(values_only=True))

        # 헤더 행(지역명이 있는 행)을 찾는다.
        header_row_idx = None
        for i, row in enumerate(all_rows):
            cells = [str(c).strip() if c else "" for c in row]
            if "서울" in cells and "경기" in cells:
                header_row_idx = i
                break

        if header_row_idx is None:
            return []

        # 헤더 행에서 지역명 → 열 인덱스 매핑을 만든다.
        header = all_rows[header_row_idx]
        region_col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(header):
            cell_str = str(cell).strip() if cell else ""
            if cell_str in self.REGION_COLS:
                region_col_map[cell_str] = col_idx
            # 헤더에서 지역 이후의 첫 번째 '계' 셀을 전국 합계로 취급한다.
            if cell_str == "계" and "전국" not in region_col_map and region_col_map:
                region_col_map["전국"] = col_idx

        # 지역별 합산 딕셔너리
        region_totals: dict[str, int] = {r: 0 for r in region_col_map}

        # 데이터 행을 순회하며 지정된 연료의 "소계/계" 행(모든 차종·용도의 공식 합계)을 읽는다.
        # ─ A열(연료)·B열(차종)이 None이면 이전 값을 유지한다 (병합 셀 대응).
        #   2018-2022 파일은 같은 연료/차종 그룹의 행을 병합 셀로 표현하므로
        #   첫 행을 제외한 나머지 행은 A·B가 None 으로 읽힌다.
        # ─ 알고리즘: A열에서 연료명을 찾은 뒤 해당 연료 섹션 끝의
        #   "소계(B열)/계(C열)" 행 1개를 직접 읽어 지역별 값을 취득한다.
        #   수소·수소전기 두 연료가 모두 존재하는 경우 각각의 소계/계 행을 합산한다.
        current_fuel    = ""
        current_vehicle = ""

        for row in all_rows[header_row_idx + 1:]:
            if not row:
                continue

            # A열: 연료 — 새 값이 있으면 갱신 및 차종 초기화
            fuel_raw = str(row[0]).strip() if row[0] else ""
            if fuel_raw:
                current_fuel    = fuel_raw
                current_vehicle = ""           # 연료 섹션이 바뀌면 차종 초기화
            fuel = current_fuel

            if not fuel:
                continue                       # 아직 연료가 확정되지 않은 초기 행

            if fuel_filter and fuel not in fuel_filter:
                continue                       # 대상 외 연료는 건너뜀

            # B열: 차종 — 새 값이 있으면 갱신 (병합 셀이면 이전 차종 유지)
            vehicle_raw = str(row[1]).strip() if row[1] else ""
            if vehicle_raw:
                current_vehicle = vehicle_raw
            vehicle = current_vehicle

            usage = str(row[2]).strip() if row[2] else ""     # C열: 용도

            # 소계/계 행만 사용: B열='소계'(전체 차종 합) + C열='계'(사업용+비사업용 합)
            # 이 행이 해당 연료 섹션의 공식 합계(지역별 등록 총수)다.
            if vehicle != "소계" or usage != "계":
                continue

            for region, col_idx in region_col_map.items():
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                # math.isnan(raw)이나 pd.isna(raw) 인지 raw == raw 로 체크
                count = int(raw) if isinstance(raw, (int, float)) and raw == raw else 0
                region_totals[region] = region_totals.get(region, 0) + count

        return [
            CarRegistrationItem(stat_year=year, region=r, count=c)
            for r, c in region_totals.items()
            if c > 0
        ]
