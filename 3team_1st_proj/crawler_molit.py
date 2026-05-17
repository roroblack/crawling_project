# 운영체제 환경변수 값을 읽기 위해 os 모듈을 가져온다.
import os
import io

# 현재 날짜와 시간을 저장하기 위해 datetime을 가져온다.
from datetime import datetime

# 엑셀 파일을 파싱하기 위해 openpyxl을 가져온다.
import openpyxl

# HTML 문서를 분석하기 위해 BeautifulSoup을 가져온다.
from bs4 import BeautifulSoup

# .env 파일에 저장된 환경변수를 읽기 위해 load_dotenv를 가져온다.
from dotenv import load_dotenv

# Playwright 비동기 API를 사용한다.
from playwright.async_api import async_playwright

# Streamlit 환경에서 asyncio 이벤트 루프를 별도 스레드에서 실행하기 위해 가져온다.
from concurrent.futures import ThreadPoolExecutor

# models.py에 정의된 CarRegistrationItem 데이터 클래스를 가져온다.
from models import CarRegistrationItem

from db import REGIONS


# .env 파일을 읽어서 os.getenv()로 사용할 수 있게 한다.
load_dotenv()


# 국토교통부 통계누리(stat.molit.go.kr)에서 자동차 등록 현황 엑셀을 내려받아
# 수소차 데이터를 추출하는 크롤러이다.
class MolitCarCrawler:
    # 10.연료별_등록현황 시트의 지역 열 이름 목록 (헤더 행 기준)
    REGION_COLS = REGIONS

    # 수집할 연료 종류 기본값 (수소차 관련만)
    TARGET_FUELS = {"수소", "수소전기"}

    # 객체가 생성될 때 자동으로 실행되는 생성자이다.
    def __init__(self):
        # 크롤링 대상 페이지 URL (자동차 등록자료 통계, hRsId=58)
        self.url = "https://stat.molit.go.kr/portal/cate/statMetaView.do?hRsId=58"

        # 엑셀 임시 저장 폴더 (프로젝트 내부, .gitignore에 추가 권장)
        self.download_dir = os.path.join(
            os.path.dirname(__file__), "molit_downloads"
        )
        os.makedirs(self.download_dir, exist_ok=True)

    # 수소차 연도별 등록현황을 수집하는 대표 메서드이다.
    # 당월 등록된 자동차 수가 전체 누계니까 연도별 12월 데이터만 수집한다. 
    # years_back년치 연말(12월) 파일 + 현재 월 파일만 내려받아 
    # 10.연료별_등록현황 시트에서 수소/수소전기 행을 추출한다.
    # 전체 월 수집(96개) 대신 연말 스냅샷(9개)만 받으므로 속도가 빠르다.
    # fuel_filter: None이면 TARGET_FUELS(수소류)만, 빈 set이면 전체 연료 수집
    def crawl(
        self,
        years_back: int = 8,
        fuel_filter: set | None = None,
    ) -> list[CarRegistrationItem]:
        
        if fuel_filter is None:
            fuel_filter = self.TARGET_FUELS

        # 페이지 HTML을 가져온 뒤 링크 목록을 한 번만 추출한다.
        html    = self._get_html()
        links   = self._parse_links(html)

        all_items: list[CarRegistrationItem] = []
        now     = datetime.now()

        # 현재 월 + 최근 years_back년의 12월(연말 스냅샷) 목록을 순회한다.
        for year, month in self._build_year_list(years_back):
            link = self._find_link(links, year, month)
            if link is None:
                continue

            # 자료를 클릭해서 xlsx_bytes 로 받는다.
            xlsx_bytes = self._get_file(link["onclick"])
            if xlsx_bytes is None:
                continue

            # xlsx_bytes 를 10.연료별_등록현황 시트 기준으로 파싱해서 CarRegistrationItem 목록을 얻는다.
            # 이후 all_items 에 누적한다.
            items = self._parse(xlsx_bytes, year, month, fuel_filter, now)
            all_items.extend(items)

        return all_items

    # 전체 차종 연도별 등록현황을 수집하는 메서드이다.
    # 최신 파일 1개만 내려받아 19.연도별 자동차 등록현황 시트를 파싱한다.
    # 파일 1개에 2007년~현재까지의 연도별 전국 합계가 모두 들어 있다.
    # stat_month=0 은 특정 월이 아닌 연간 합계를 의미한다.
    def crawl_yearly(self) -> list[CarRegistrationItem]:
        # 연도별로 자동차등록현황 xlsx 파일을 크롤링
        html = self._get_html()
        links = self._parse_links(html)

        # 자동차 등록자료 중 가장 첫 번째(최신) 파일을 가져온다.
        car_links = [li for li in links if "자동차 등록자료" in li.get("text", "")]
        if not car_links:
            return []

        # 다운로드한 xlsx 파일의 용량이 0이면 파일이 없거나 못받은 거니까 [] 을 반환한다.
        xlsx_bytes = self._get_file(car_links[0]["onclick"])
        if xlsx_bytes is None:
            return []

        # xlsx 파일 읽기에 성공시 읽어온 데이터를 반환한다.
        return self._parse_yearly(xlsx_bytes, datetime.now())

    # ── 내부 헬퍼 메서드들 ──────────────────────────────────────────────

    def _build_year_list(self, years_back: int) -> list[tuple[int, int]]:
        # 현재 월 + 최근 years_back년의 12월(연말)로 이루어진 목록을 만든다.
        # 예) years_back=8, 현재=2026-04 → [(2026,4),(2025,12),...,(2018,12)]
        now     = datetime.now()
        result  = [(now.year, now.month)]
        for i in range(1, years_back + 1):
            result.append((now.year - i, 12))
        return result

    def _get_html(self) -> str:
        # 웹페이지 HTML을 가져온다.
        # Streamlit은 asyncio 이벤트 루프를 사용하기 때문에
        # 별도의 스레드에서 실행하면 NotImplementedError를 피할 수 있다.
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._run_playwright)
            return future.result()

    def _run_playwright(self) -> str:
        # Windows에서 ProactorEventLoop을 명시적으로 생성하여
        # Streamlit/tornado의 SelectorEventLoop 정책과의 충돌을 피한다.
        import asyncio
        loop = asyncio.ProactorEventLoop()
        return loop.run_until_complete(self._async_get_html())

    # async_playwright를 사용해 비동기로 HTML을 가져온다.
    # 기다리는 페이지를 기다리는 동안 Streamlit UI가 멈추지 않게 asyncio 이벤트 루프를 활용한다.
    async def _async_get_html(self) -> str:
        # async_playwright를 사용해 비동기로 HTML을 가져온다.
        async with async_playwright() as p:
            browser     = await p.chromium.launch(headless=True)
            page        = await browser.new_page(
                user_agent=(
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 Chrome/120.0 Safari/537.36"
                )
            )
            await page.goto(self.url, wait_until="networkidle", timeout=60000)
            await page.wait_for_timeout(1500)
            html = await page.content()
            await browser.close()
        return html

    def _parse_links(self, html: str) -> list[dict]:
        # HTML에서 downFile 링크 정보를 추출한다.
        # 각 항목은 {"text": 표시 텍스트, "onclick": onclick 속성값} 형태이다.
        soup    = BeautifulSoup(html, "lxml")
        links   = []
        for a in soup.select("a[onclick*='downFile']"):
            links.append({
                "text": a.get_text(strip=True),
                "onclick": a.get("onclick", ""),
            })
        return links

    def _find_link(self, links: list[dict], year: int, month: int) -> dict | None:
        # 링크 텍스트에서 연도와 월이 일치하는 자동차 등록자료 항목을 찾는다.
        # 파일명 패턴 예시:
        #   "2026년 4월 자동차 등록자료 통계.xlsx"
        #   "2026년 04월 자동차 등록자료 통계.xlsx"  (0 패딩 있는 경우)
        patterns = [
            f"{year}년 {month}월 자동차 등록자료",
            f"{year}년 {month:02d}월 자동차 등록자료",
        ]
        for link in links:
            text = link.get("text", "")
            if any(p in text for p in patterns):
                return link
        return None

    def _get_file(self, onclick: str) -> bytes | None:
        # onclick 속성값을 전달해 Playwright가 해당 링크를 클릭 후 파일을 내려받는다.
        with ThreadPoolExecutor(max_workers=1) as executor:
            future = executor.submit(self._run_download, onclick)
            return future.result()

    def _run_download(self, onclick: str) -> bytes | None:
        import asyncio
        loop = asyncio.ProactorEventLoop()
        return loop.run_until_complete(self._async_download(onclick))

    async def _async_download(self, onclick: str) -> bytes | None:
        async with async_playwright() as p:
            browser = await p.chromium.launch(headless=True)
            ctx = await browser.new_context(accept_downloads=True)
            page = await ctx.new_page()
            await page.goto(self.url, wait_until="networkidle", timeout=60000)

            try:
                # onclick 속성이 동일한 링크를 클릭해 다운로드 이벤트를 유도한다.
                async with page.expect_download(timeout=30000) as dl_info:
                    await page.eval_on_selector_all(
                        "a[onclick*='downFile']",
                        "(els, oc) => { for (const e of els) { if (e.getAttribute('onclick') === oc) { e.click(); break; } } }",
                        onclick,
                    )
                download = await dl_info.value

                # 다운로드된 파일을 bytes로 읽는다 (디스크 저장 없이 메모리로 처리).
                path = await download.path()
                with open(path, "rb") as f:
                    data = f.read()
            except Exception:
                data = None
            finally:
                await browser.close()

        return data

    def _parse(
        self,
        xlsx_bytes: bytes,
        year: int,
        month: int,
        fuel_filter: set,
        crawled_at: datetime,
    ) -> list[CarRegistrationItem]:
        # 10.연료별_등록현황 시트를 파싱해 CarRegistrationItem 목록을 반환한다.
        # 지역(열)을 행으로 언피벗하여 region 컬럼으로 저장한다.
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
        if "10.연료별_등록현황" not in wb.sheetnames:
            return []

        ws = wb["10.연료별_등록현황"]
        items: list[CarRegistrationItem] = []
        all_rows = list(ws.iter_rows(values_only=True))

        # 헤더 행(지역명이 있는 행)을 찾는다.
        header_row_idx = None
        for i, row in enumerate(all_rows):
            cells = [str(c).strip() if c else "" for c in row]
            if "서울" in cells and "경기" in cells:
                header_row_idx = i
                break

        if header_row_idx is None:
            return []

        # 헤더 행에서 지역명 → 열 인덱스 매핑을 만든다.
        header = all_rows[header_row_idx]
        region_col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(header):
            cell_str = str(cell).strip() if cell else ""
            if cell_str in self.REGION_COLS:
                region_col_map[cell_str] = col_idx
            # 헤더에서 지역 이후의 첫 번째 '계' 셀을 전국 합계로 취급한다.
            if cell_str == "계" and "전국" not in region_col_map and region_col_map:
                region_col_map["전국"] = col_idx

        # 데이터 행을 순회하며 지정된 연료 행만 언피벗하여 추출한다.
        for row in all_rows[header_row_idx + 1:]:
            if not row or not row[0]:
                continue

            fuel = str(row[0]).strip()       # A열: 연료 종류
            if fuel_filter and fuel not in fuel_filter:
                continue

            vehicle = str(row[1]).strip() if row[1] else ""  # B열: 차종
            usage = str(row[2]).strip() if row[2] else ""    # C열: 용도

            # 차종 또는 용도가 비어 있으면 소제목 행이므로 건너뛴다.
            if not vehicle or not usage:
                continue

            for region, col_idx in region_col_map.items():
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                count = int(raw) if isinstance(raw, (int, float)) and raw == raw else 0

                items.append(
                    CarRegistrationItem(
                        stat_year=year,
                        stat_month=month,
                        fuel_type=fuel,
                        vehicle_type=vehicle,
                        usage_type=usage,
                        region=region,
                        count=count,
                        crawled_at=crawled_at,
                    )
                )

        return items

    def _parse_yearly(
        self,
        xlsx_bytes: bytes,
        crawled_at: datetime,
    ) -> list[CarRegistrationItem]:
        # 19.연도별 자동차 등록현황 시트를 파싱한다.
        # 이 시트 1개에 2007년~현재까지의 연도별 전국 합계가 모두 있다.
        # fuel_type="합계", region="전국", stat_month=0(연간 합계) 으로 저장한다.
        wb = openpyxl.load_workbook(
            io.BytesIO(xlsx_bytes), read_only=True, data_only=True
        )
        if "19.연도별 자동차 등록현황" not in wb.sheetnames:
            return []

        ws = wb["19.연도별 자동차 등록현황"]
        items: list[CarRegistrationItem] = []
        all_rows = list(ws.iter_rows(values_only=True))

        # 용도 헤더 행("년도"가 첫 번째 셀인 행)을 찾는다.
        usage_row_idx = None
        for i, row in enumerate(all_rows):
            if row and str(row[0]).strip() == "년도":
                usage_row_idx = i
                break

        if usage_row_idx is None:
            return []

        # 차종 헤더 행은 용도 헤더 행의 바로 위 행이다.
        vehicle_row = all_rows[usage_row_idx - 1]
        usage_row = all_rows[usage_row_idx]

        # 열 인덱스 → (vehicle_type, usage_type) 매핑을 구성한다.
        # 차종 행에서 앞 채우기(forward-fill)로 각 열의 차종을 결정한다.
        col_map: dict[int, tuple[str, str]] = {}
        current_vehicle = None
        for col_idx in range(1, len(usage_row)):
            v = str(vehicle_row[col_idx]).strip() if col_idx < len(vehicle_row) and vehicle_row[col_idx] else ""
            if v:
                current_vehicle = v
            u = str(usage_row[col_idx]).strip() if usage_row[col_idx] else ""
            if current_vehicle and u:
                col_map[col_idx] = (current_vehicle, u)

        # 데이터 행을 순회하며 연도별 수치를 추출한다.
        for row in all_rows[usage_row_idx + 1:]:
            if not row or not row[0]:
                continue

            # 첫 번째 셀이 4자리 연도 숫자인 행만 처리한다.
            year_val = row[0]
            if not isinstance(year_val, (int, float)) or int(year_val) < 2000:
                continue

            stat_year = int(year_val)

            for col_idx, (vehicle_type, usage_type) in col_map.items():
                if col_idx >= len(row):
                    continue
                raw = row[col_idx]
                count = int(raw) if isinstance(raw, (int, float)) and raw == raw else 0

                items.append(
                    CarRegistrationItem(
                        stat_year=stat_year,
                        stat_month=0,        # 0 = 연간 합계 (특정 월 아님)
                        fuel_type="합계",    # 연료 구분 없는 전체 합계
                        vehicle_type=vehicle_type,
                        usage_type=usage_type,
                        region="전국",       # 지역 구분 없는 전국 합계
                        count=count,
                        crawled_at=crawled_at,
                    )
                )

        return items
